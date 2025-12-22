"""
Backend функция для автоматической отправки email-отчетов по аналитике
Генерирует Excel отчет и отправляет на указанный email
"""
import json
import os
from typing import Dict, Any
from datetime import datetime, timedelta
import psycopg2
from psycopg2.extras import RealDictCursor
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import io
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    openpyxl = None

PROGRAM_NAMES = {
    'family': 'Семейная ипотека',
    'it': 'IT ипотека',
    'military': 'Военная ипотека',
    'rural': 'Сельская ипотека',
    'basic': 'Базовая ипотека',
    'unknown': 'Помочь подобрать'
}

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, X-Admin-Password',
                'Access-Control-Max-Age': '86400'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    if method == 'POST':
        # Check admin password
        headers = event.get('headers', {})
        admin_password = headers.get('X-Admin-Password') or headers.get('x-admin-password')
        expected_password = os.environ.get('ADMIN_PASSWORD', 'admin123')
        
        if admin_password != expected_password:
            return {
                'statusCode': 401,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Content-Type': 'application/json'
                },
                'body': json.dumps({'error': 'Unauthorized'}),
                'isBase64Encoded': False
            }
        
        body_data = json.loads(event.get('body', '{}'))
        recipient_email = body_data.get('email')
        days = int(body_data.get('days', 7))
        
        if not recipient_email:
            return {
                'statusCode': 400,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Content-Type': 'application/json'
                },
                'body': json.dumps({'error': 'Email is required'}),
                'isBase64Encoded': False
            }
        
        # Get analytics data
        dsn = os.environ.get('DATABASE_URL')
        if not dsn:
            return {
                'statusCode': 500,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Content-Type': 'application/json'
                },
                'body': json.dumps({'error': 'Database not configured'}),
                'isBase64Encoded': False
            }
        
        conn = psycopg2.connect(dsn)
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        try:
            # Get daily views
            cur.execute("""
                SELECT 
                    DATE(created_at) as date,
                    COUNT(*) as views
                FROM analytics_events
                WHERE event_type = 'page_view' 
                    AND created_at >= CURRENT_DATE - INTERVAL '%s days'
                GROUP BY DATE(created_at)
                ORDER BY date DESC
            """ % days)
            daily_views = cur.fetchall()
            
            # Get daily applications
            cur.execute("""
                SELECT 
                    DATE(created_at) as date,
                    COUNT(*) as applications
                FROM analytics_events
                WHERE event_type = 'application_sent'
                    AND created_at >= CURRENT_DATE - INTERVAL '%s days'
                GROUP BY DATE(created_at)
                ORDER BY date DESC
            """ % days)
            daily_applications = cur.fetchall()
            
            # Get traffic sources
            cur.execute("""
                SELECT 
                    COALESCE(source, 'Прямой переход') as source,
                    COUNT(*) as count
                FROM analytics_events
                WHERE event_type = 'page_view'
                    AND created_at >= CURRENT_DATE - INTERVAL '%s days'
                GROUP BY source
                ORDER BY count DESC
                LIMIT 10
            """ % days)
            traffic_sources = cur.fetchall()
            
            # Get popular programs
            cur.execute("""
                SELECT 
                    program,
                    COUNT(*) as count
                FROM analytics_events
                WHERE event_type = 'application_sent'
                    AND program IS NOT NULL
                    AND created_at >= CURRENT_DATE - INTERVAL '%s days'
                GROUP BY program
                ORDER BY count DESC
            """ % days)
            popular_programs = cur.fetchall()
            
            # Get totals
            cur.execute("""
                SELECT 
                    COUNT(CASE WHEN event_type = 'page_view' THEN 1 END) as total_views,
                    COUNT(CASE WHEN event_type = 'application_sent' THEN 1 END) as total_applications
                FROM analytics_events
                WHERE created_at >= CURRENT_DATE - INTERVAL '%s days'
            """ % days)
            totals = cur.fetchone()
            
        finally:
            cur.close()
            conn.close()
        
        # Generate Excel report
        if not openpyxl:
            return {
                'statusCode': 500,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Content-Type': 'application/json'
                },
                'body': json.dumps({'error': 'openpyxl not available'}),
                'isBase64Encoded': False
            }
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Сводка"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ws_summary['A1'] = 'Отчет по аналитике сайта'
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A2'] = 'Период'
        ws_summary['B2'] = f'{days} дней'
        ws_summary['A3'] = 'Дата создания'
        ws_summary['B3'] = datetime.now().strftime('%d.%m.%Y %H:%M')
        
        ws_summary['A5'] = 'Метрика'
        ws_summary['B5'] = 'Значение'
        ws_summary['A5'].fill = header_fill
        ws_summary['B5'].fill = header_fill
        ws_summary['A5'].font = header_font
        ws_summary['B5'].font = header_font
        
        ws_summary['A6'] = 'Всего просмотров'
        ws_summary['B6'] = totals['total_views']
        ws_summary['A7'] = 'Всего заявок'
        ws_summary['B7'] = totals['total_applications']
        ws_summary['A8'] = 'Конверсия'
        conversion = (totals['total_applications'] / totals['total_views'] * 100) if totals['total_views'] > 0 else 0
        ws_summary['B8'] = f"{conversion:.2f}%"
        
        # Daily views sheet
        ws_views = wb.create_sheet("Просмотры по дням")
        ws_views['A1'] = 'Дата'
        ws_views['B1'] = 'Просмотры'
        ws_views['A1'].fill = header_fill
        ws_views['B1'].fill = header_fill
        ws_views['A1'].font = header_font
        ws_views['B1'].font = header_font
        
        for idx, row in enumerate(daily_views, start=2):
            ws_views[f'A{idx}'] = row['date'].strftime('%d.%m.%Y') if hasattr(row['date'], 'strftime') else str(row['date'])
            ws_views[f'B{idx}'] = row['views']
        
        # Daily applications sheet
        ws_apps = wb.create_sheet("Заявки по дням")
        ws_apps['A1'] = 'Дата'
        ws_apps['B1'] = 'Заявки'
        ws_apps['A1'].fill = header_fill
        ws_apps['B1'].fill = header_fill
        ws_apps['A1'].font = header_font
        ws_apps['B1'].font = header_font
        
        for idx, row in enumerate(daily_applications, start=2):
            ws_apps[f'A{idx}'] = row['date'].strftime('%d.%m.%Y') if hasattr(row['date'], 'strftime') else str(row['date'])
            ws_apps[f'B{idx}'] = row['applications']
        
        # Traffic sources sheet
        ws_sources = wb.create_sheet("Источники трафика")
        ws_sources['A1'] = 'Источник'
        ws_sources['B1'] = 'Количество'
        ws_sources['C1'] = 'Процент'
        for cell in ['A1', 'B1', 'C1']:
            ws_sources[cell].fill = header_fill
            ws_sources[cell].font = header_font
        
        total_sources = sum(s['count'] for s in traffic_sources)
        for idx, row in enumerate(traffic_sources, start=2):
            ws_sources[f'A{idx}'] = row['source']
            ws_sources[f'B{idx}'] = row['count']
            percentage = (row['count'] / total_sources * 100) if total_sources > 0 else 0
            ws_sources[f'C{idx}'] = f"{percentage:.1f}%"
        
        # Popular programs sheet
        ws_programs = wb.create_sheet("Популярные программы")
        ws_programs['A1'] = 'Программа'
        ws_programs['B1'] = 'Заявок'
        ws_programs['C1'] = 'Процент'
        for cell in ['A1', 'B1', 'C1']:
            ws_programs[cell].fill = header_fill
            ws_programs[cell].font = header_font
        
        total_programs = sum(p['count'] for p in popular_programs)
        for idx, row in enumerate(popular_programs, start=2):
            ws_programs[f'A{idx}'] = PROGRAM_NAMES.get(row['program'], row['program'])
            ws_programs[f'B{idx}'] = row['count']
            percentage = (row['count'] / total_programs * 100) if total_programs > 0 else 0
            ws_programs[f'C{idx}'] = f"{percentage:.1f}%"
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Send email
        smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
        smtp_port = int(os.environ.get('SMTP_PORT', '587'))
        smtp_user = os.environ.get('SMTP_USER')
        smtp_password = os.environ.get('SMTP_PASSWORD')
        
        if not smtp_user or not smtp_password:
            return {
                'statusCode': 500,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Content-Type': 'application/json'
                },
                'body': json.dumps({'error': 'SMTP credentials not configured'}),
                'isBase64Encoded': False
            }
        
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = recipient_email
        msg['Subject'] = f'Отчет по аналитике сайта за {days} дней'
        
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2>Отчет по аналитике сайта</h2>
            <p>Период: <strong>{days} дней</strong></p>
            <p>Дата создания: <strong>{datetime.now().strftime('%d.%m.%Y %H:%M')}</strong></p>
            
            <h3>Основные показатели:</h3>
            <ul>
                <li>Всего просмотров: <strong>{totals['total_views']}</strong></li>
                <li>Всего заявок: <strong>{totals['total_applications']}</strong></li>
                <li>Конверсия: <strong>{conversion:.2f}%</strong></li>
            </ul>
            
            <p>Детальная статистика во вложенном Excel-файле.</p>
            
            <p style="color: #666; font-size: 12px;">
                Автоматический отчет с сайта ипотекакрым.рф
            </p>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        # Attach Excel file
        filename = f'Аналитика_{days}дней_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(excel_buffer.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)
        
        # Send email
        try:
            server = smtplib.SMTP(smtp_host, smtp_port)
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
            server.quit()
            
            return {
                'statusCode': 200,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Content-Type': 'application/json'
                },
                'body': json.dumps({'message': 'Report sent successfully', 'email': recipient_email}),
                'isBase64Encoded': False
            }
        except Exception as e:
            return {
                'statusCode': 500,
                'headers': {
                    'Access-Control-Allow-Origin': '*',
                    'Content-Type': 'application/json'
                },
                'body': json.dumps({'error': f'Failed to send email: {str(e)}'}),
                'isBase64Encoded': False
            }
    
    return {
        'statusCode': 405,
        'headers': {
            'Access-Control-Allow-Origin': '*',
            'Content-Type': 'application/json'
        },
        'body': json.dumps({'error': 'Method not allowed'}),
        'isBase64Encoded': False
    }